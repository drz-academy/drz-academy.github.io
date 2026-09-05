#!/usr/bin/env python3
import json
import csv
import pandas as pd
import os
import argparse
import sys

# Agregar club/bin al path para poder importar config si se corre desde otro lado
sys.path.append(os.path.join(os.path.dirname(__file__)))
import config

import datetime

# Rutas por defecto asumiendo que se corre desde el root del repositorio
CLUB_DIR = "club"
PERSONAL_DIR = os.path.join(CLUB_DIR, "personal")
CLUB_DRZ_DRIVE = os.path.normpath(os.path.join(CLUB_DIR, config.CLUB_DRZ_DRIVE))
INFO_DIR = config.get_info_dir(CLUB_DIR)

MEMBERS_JSON = os.path.join(INFO_DIR, "drz-club-members.json")
CERTIFICADOS_CSV = os.path.join(PERSONAL_DIR, "certificados.csv")
OUTPUT_EXCEL = os.path.join(INFO_DIR, "drz-club-members.xlsx")

CURSOS_MAPPING = {
    'Cuántica a Pie': ['cuantica_a_pie', 'cuantica_a_pie_permanente'],
    'Catastrofísica': ['catastrofisica'],
    'Einstein Relativamente Fácil': ['einstein'],
    'Mundo Cuántico': ['mundo_cuantico'],
    'El Rompecabezas de la Materia': ['rompecabezas_materia'],
    'Astropython': ['astropython'],
    'Python para el fin del mundo': ['python_fin_mundo'],
    'Master Class Extraterrestre': ['masterclass_extraterrestre']
}

def load_data(members_path, certs_path):
    with open(members_path, 'r', encoding='utf-8') as f:
        members = json.load(f)
    
    certificados = []
    if os.path.exists(certs_path):
        with open(certs_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                certificados.append(row)
    else:
        print(f"Advertencia: No se encontró el archivo de certificados en {certs_path}")
            
    return members, certificados

def main():
    parser = argparse.ArgumentParser(description="Genera documento maestro en Excel de los miembros del Club.")
    parser.add_argument('--out', default=OUTPUT_EXCEL, help='Ruta del archivo Excel de salida')
    args = parser.parse_args()

    members, certificados = load_data(MEMBERS_JSON, CERTIFICADOS_CSV)
    
    # Map certificates by document or email to course_ids
    certs_by_doc = {}
    certs_by_email = {}
    
    for c in certificados:
        doc = c.get('documento', '').strip()
        correo = c.get('correo', '').strip().lower()
        url = c.get('url', '').strip()
        cid = c.get('curso_id', '').strip()
        
        if doc:
            if doc not in certs_by_doc: certs_by_doc[doc] = {}
            certs_by_doc[doc][cid] = url
        if correo:
            if correo not in certs_by_email: certs_by_email[correo] = {}
            certs_by_email[correo][cid] = url

    # Prepare data for DataFrame
    rows = []
    cursos_list = list(CURSOS_MAPPING.keys())
    
    for m in members:
        doc = str(m.get('documento', '')).strip()
        correo = str(m.get('correo', '')).strip().lower()
        
        # Get this person's certificates
        person_certs = {}
        if doc and doc in certs_by_doc:
            person_certs.update(certs_by_doc[doc])
        if correo and correo in certs_by_email:
            person_certs.update(certs_by_email[correo])
            
        cursos_asistidos = m.get('cursos_participados', [])
        
        row = {
            'Nombre': m.get('nombre', ''),
            'Cédula': doc,
            'Correo electrónico': m.get('correo', ''),
            'Número de cursos asistidos': len(cursos_asistidos),
            'Categoría': m.get('categoria', ''),
        }
        
        certificados_count = 0
        
        for curso_nombre in cursos_list:
            cids = CURSOS_MAPPING[curso_nombre]
            
            # check if certified
            cert_url = None
            for cid in cids:
                if cid in person_certs:
                    cert_url = person_certs[cid]
                    break
            
            if cert_url:
                row[curso_nombre] = cert_url 
                certificados_count += 1
            elif curso_nombre in cursos_asistidos:
                row[curso_nombre] = 'Asistió'
            else:
                row[curso_nombre] = 'No'
                
        row['Número de cursos certificados'] = certificados_count
        
        rows.append(row)
        
    df = pd.DataFrame(rows)
    
    # Reorder columns
    cols = ['Nombre', 'Cédula', 'Correo electrónico', 'Número de cursos asistidos', 'Número de cursos certificados', 'Categoría'] + cursos_list
    df = df[cols]
    
    # Write to Excel
    writer = pd.ExcelWriter(args.out, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Maestro', startrow=1)
    
    # Convert URLs to actual hyperlinks in Excel
    workbook = writer.book
    worksheet = writer.sheets['Maestro']
    
    # Añadir encabezado con la última actualización
    today = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    worksheet['A1'] = f"Última actualización: {today}"
    worksheet.merge_cells('A1:E1')
    
    for row_idx in range(3, len(df) + 3):
        for col_idx, col_name in enumerate(cols, start=1):
            if col_name in cursos_list:
                cell_value = df.iloc[row_idx-3][col_name]
                if str(cell_value).startswith('http'):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.hyperlink = cell_value
                    cell.value = 'Certificado'
                    cell.style = "Hyperlink"
                    
    # Adjust column widths for better readability
    from openpyxl.utils import get_column_letter
    for col_idx, col in enumerate(worksheet.columns, start=1):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = min(adjusted_width, 40)
                    
    writer.close()
    print(f"Documento maestro creado exitosamente en {args.out}")

if __name__ == '__main__':
    main()
